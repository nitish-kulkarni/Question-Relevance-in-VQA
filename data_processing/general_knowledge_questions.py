import openpyxl
import sys

### Usage
### python general_knowledge_questions.py
### This function extracts all general knowledge questions from genericqdata.xlsx

wb=openpyxl.load_workbook('../data/visual_vs_nonvisual/genericqdata.xlsx')

sheets=wb.get_sheet_names()
questiondata=[]
for sheet in sheets:
	sheet_name=wb.get_sheet_by_name(sheet)
	numdata=sheet_name.max_row
	for i in range(1,numdata):
		if sheet_name.cell(row=i,column=2).value!=None:
			questiondata.append([sheet_name.cell(row=i,column=2).value,1])

for q in questiondata:
	sys.stdout.write(q[0].encode('utf-8').strip() + "\n")